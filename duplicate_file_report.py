import hashlib
import os
import pandas as pd
from typing import List, Tuple
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from collections import defaultdict
import re

def calculate_full_hash(file_path: str, chunk_size: int = 1024 * 1024) -> str:
    """Calculate the full MD5 hash of a file."""
    hash_algo = hashlib.md5()
    with open(file_path, 'rb') as f:
        while chunk := f.read(chunk_size):
            hash_algo.update(chunk)
    return hash_algo.hexdigest()

def is_form_type(filename: str) -> str:
    """Check if the file name follows the potential form type pattern."""
    match = re.match(r"^\([A-Za-z0-9]+\)", filename)
    return "Form(maybe)" if match else ""

def list_all_files(folder_path: str, folder_label: str) -> List[Tuple[str, str, str, int, str, str]]:
    """List all files in a folder with details: file name, hash, type, size, folder label, and content type."""
    file_list = []
    for root, _, files in os.walk(folder_path):
        for file in files:
            file_path = os.path.join(root, file)
            file_size = os.path.getsize(file_path)
            file_hash = calculate_full_hash(file_path)
            file_extension = os.path.splitext(file)[1].lower()  # Get file extension
            content_type = is_form_type(file)  # Check for form type
            
            # Append file details: file name, hash, type, size, folder label, and content type
            file_list.append((os.path.basename(file), file_hash, file_extension, file_size, folder_label, content_type))
    return file_list

def generate_styled_excel_report(file_list: List[Tuple[str, str, str, int, str, str]], report_path: str) -> None:
    """Generate a styled Excel report listing all files from both folders, sorted by file size."""
    # Convert data to a DataFrame
    df = pd.DataFrame(file_list, columns=[
        "File Name", "File Hash", "File Type", "File Size", "Folder", "Content Type"
    ])
    
    # Sort by file size in ascending order
    df.sort_values(by="File Size", inplace=True)

    # Find duplicate hashes for highlighting
    hash_counts = defaultdict(int)
    for file_hash in df["File Hash"]:
        hash_counts[file_hash] += 1

    # Create an Excel workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "File List Report"
    
    # Define styles
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold for header
    folder1_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light blue for Folder1 name
    folder2_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")  # Light pink for Folder2 name
    duplicate_hash_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow for duplicate hashes
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    headers = ["File Name", "File Hash", "File Type", "File Size", "Folder", "Content Type"]
    ws.append(headers)
    for cell in ws[1]:  # Apply header style
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = center_alignment
    
    # Write data rows
    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)
        row_num = ws.max_row
        folder_name = row[4]
        file_hash = row[1]

        # Apply color based on folder name
        folder_cell = ws[f"E{row_num}"]
        if folder_name == "Folder1":
            folder_cell.fill = folder1_fill
        elif folder_name == "Folder2":
            folder_cell.fill = folder2_fill
        
        # Highlight duplicate hash cells
        hash_cell = ws[f"B{row_num}"]
        if hash_counts[file_hash] > 1:
            hash_cell.fill = duplicate_hash_fill
        
        # Center alignment for all cells
        for cell in ws[row_num]:
            cell.alignment = center_alignment

    # Adjust column width for better readability
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # Save the workbook
    wb.save(report_path)

# Main execution
if __name__ == "__main__":
    folder1 = "folder1"  # Replace with the path to the first folder
    folder2 = "folder2"  # Replace with the path to the second folder
    report_path = "all_files_report.xlsx"  # Path to save the Excel report
    
    # List all files from both folders with labels
    all_files = list_all_files(folder1, "Folder1") + list_all_files(folder2, "Folder2")
    generate_styled_excel_report(all_files, report_path)
    
    print(f"Styled Excel report generated: {report_path}")
